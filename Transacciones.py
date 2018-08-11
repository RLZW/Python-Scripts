from bs4 import BeautifulSoup
import re
import openpyxl
import time
import xlsxwriter
start_time = time.time()

Archivo = 'Cuentas.xlsx'
Hoja = 'Sheet1'
html = "BancaNet_Citibanamex.com.html"
wb = openpyxl.load_workbook(Archivo)
sheet = wb.get_sheet_by_name(Hoja)
with open(html) as fp:
    soup = BeautifulSoup(fp)
persona = []

def main():
    #Cambia las cuentas a diccionario
    cuentas = accountToDict(createListOfAccounts())

    #Divs Información de transferencias.
    DivTransacciones = getDiv("cbolui-table-wrapper cbolui-clearfix")
    Transacciones = []

    #transaccion = "'\n\n\n\n\n\n\n\n\n\n\n\n\t\t\t\t\t\t\t\t\tReferencia alfanumï¿½rica:\n\t\t\t\t\t\t\t\t\n\nOrigen\nDestino\n\n\nORIGEN\nTipo:\nCHQ\nTDD\n\n\n\xa0\nSucursal:\n505\nN/A\n\n\nReferencia numï¿½rica:\n605674\nCuenta:\n69301\n61799\n\n\nFecha de autorizaciï¿½n:\n31 Jul 2018 - 20:10\nTipo de instrumento:\nN/A\nN/A\n\n\nTipo de cambio:\n$ 0,000000\nImporte:\n$ 1,800.00\n$ 1,800.00\n\n\nServicio:\nMOVIL\n\xa0\n\xa0\n\xa0\n\n\n\n\n\n\n\n\nDatos adicionales\n\xa0\nDatos de cancelaciï¿½n\n\xa0\n\n\nBanco:\nN/A\nServicio / Medio:\nN/A\n\n\nClave de rastreo:\nN/A\nSucursal:\nN/A\n\n\nInstrucciï¿½n aplicada:\n20:10:06\nFecha:\nN/A\n\n\nInstrucciï¿½n liquidada:\nN/A\nOperador:\nN/A\n\n\nCausa de Devoluciï¿½n:\nN/A\n\xa0\n\xa0\n\n\nID de ejecutivo 1:\nN/A\n\xa0\n\xa0\n\n\nNombre de ejecutivo 1:\nN/A\n\xa0\n\xa0\n\n\nID de ejecutivo 2:\nN/A\n\xa0\n\xa0\n\n\nNombre de ejecutivo 2:\nN/A\n\xa0\n\xa0\n\n\n\n'"
    for transaccion in DivTransacciones:
        #Regex para filtrar datos de la transacción 
        Importe = re.search(r'Importe:[\r\n]+\$[^"]+\$',transaccion).group() #Importe:\n$ 1,800.00\n$       
        Fecha = re.search(r'[0-9]+[^"]+',re.search(r'Fecha de autorizaciï¿½n:[\r\n]+[0-9]+[\s]+[a-zA-Z]+[\s]+[0-9]+', transaccion).group()).group().rstrip()
        Referencia = re.search(r'[\r\n][^"]+',re.search(r'Referencia numï¿½rica:[\r\n][a-zA-Z0-9//]+[\r\n]', transaccion).group().rstrip()).group().lstrip()
        Importe = re.search(r'[\d]+[^"]+[\r\n]+',re.search(r'Importe:[\r\n]+\$[^"]+\$',transaccion).group()).group().rstrip()
        CuentaOrigen = re.search(r'[\r\n]+[^"]+[\r\n]+',re.search(r'Cuenta:[\r\n]+[a-zA-Z0-9//]+[\r\n]+[a-zA-Z0-9//]+', transaccion).group()).group().rstrip().lstrip()
        CuentaDestino = re.search(r'[\r\n]+[a-zA-Z0-9//]+',re.search(r'[\r\n]+[^"]+',re.search(r'Cuenta:[\r\n]+[a-zA-Z0-9//]+[\r\n]+[a-zA-Z0-9//]+', transaccion).group()).group().rstrip().lstrip()).group().rstrip().lstrip()
        if(CuentaOrigen == 'N/A' and CuentaDestino == 'N/A'):
            Transaccion = [Fecha,Referencia,Importe,CuentaOrigen,CuentaDestino,cuentas[int(CuentaOrigen[-3:])],"NO EXISTE EN EL DICCIONARIO DE CUENTAS"]
        elif(CuentaOrigen == 'N/A'):
            Transaccion = [Fecha,Referencia,Importe,CuentaOrigen,CuentaDestino,"NO EXISTE EN EL DICCIONARIO DE CUENTAS","NO EXISTE EN EL DICCIONARIO DE CUENTAS"]
        elif(CuentaDestino == 'N/A'):
             Transaccion = [Fecha,Referencia,Importe,CuentaOrigen,CuentaDestino,cuentas[int(CuentaOrigen[-3:])],"NO EXISTE EN EL DICCIONARIO DE CUENTAS"]
        elif((int(CuentaOrigen[-3:]) in cuentas) and (int(CuentaDestino[-3:]) in cuentas)):
            Transaccion = [Fecha,Referencia,Importe,CuentaOrigen,CuentaDestino,cuentas[int(CuentaOrigen[-3:])],cuentas[int(CuentaDestino[-3:])]]
        elif(not(int(CuentaOrigen[-3:]) in cuentas) and not(int(CuentaDestino[-3:]) in cuentas)):
            Transaccion = [Fecha,Referencia,Importe,CuentaOrigen,CuentaDestino,"NO EXISTE EN EL DICCIONARIO DE CUENTAS","NO EXISTE EN EL DICCIONARIO DE CUENTAS"]
        elif(not(int(CuentaOrigen[-3:]) in cuentas)):
            Transaccion = [Fecha,Referencia,Importe,CuentaOrigen,CuentaDestino,"NO EXISTE EN EL DICCIONARIO DE CUENTAS",cuentas[int(CuentaDestino[-3:])]]
        elif(not(int(CuentaDestino[-3:]) in cuentas)): 
            Transaccion = [Fecha,Referencia,Importe,CuentaOrigen,CuentaDestino,cuentas[int(CuentaOrigen[-3:])],"NO EXISTE EN EL DICCIONARIO DE CUENTAS"]
        
        Transacciones.append(Transaccion)
        print(Transaccion)
    workbook = xlsxwriter.Workbook('Transacciones.xlsx')
    worksheet = workbook.add_worksheet()
    row = 0
    col = 0

    for Fecha,Autorización,Monto,Cuenta_Origen,Cuenta_Destino,Origen,Destino in (Transacciones):
        worksheet.write(row, col, Fecha)
        worksheet.write(row, col +1, Autorización)
        worksheet.write(row, col +2, Monto)
        worksheet.write(row, col +3, Cuenta_Origen)
        worksheet.write(row, col +4, Cuenta_Destino)
        worksheet.write(row, col +5, Origen)
        worksheet.write(row, col +6, Destino)
        row += 1
    workbook.close()
    

    print("El programa termino en: --- %s segundos ---" % (time.time() - start_time))
    
##################################################################



def getSizeOfColumn(column):
    contador = 0
    column = sheet[column]
    for _ in column:
        contador += 1
    return contador

def createListOfAccounts():
    cuentas = []
    for row in sheet.iter_rows(min_row=1, max_col=2, max_row=getSizeOfColumn("A")):
            for cell in row:
                cuentas.append(cell.value)
    print("La lista de cuentas sucia fue creada")
    return cuentas


def accountToDict(list):
    cuentas = []
    personas = []
    for i in range(len(list)):
        if i % 2 == 0:
            personas.append(list[i-1])
        else:
            cuentas.append(list[i-3])
    dictionary = dict(zip(personas,cuentas))

    print("Las cuentas fueron asiganadas a un diccionario")
    print("El diccionario es el siguiente: ")
    print(dictionary)
    print("Hay un total de "+str(len(dictionary))+" cuentas.")
    return dictionary

#############################################################

def getTd(searchText):
    temp = []
    for content in soup.find_all("td", class_=searchText):
        temp.append(content.text)
    return temp

def getDiv(searchText):
    temp = []
    for content in soup.find_all("div", class_=searchText):
        temp.append(content.text)
    return temp

def write_matrix_to_textfile(a_matrix, file_to_write):

    def compile_row_string(a_row):
        return str(a_row).strip(']').strip('[').replace(' ','')

    with open(file_to_write, 'w') as f:
        for row in a_matrix:
            f.write(compile_row_string(row)+'\n')

    return True

main()
